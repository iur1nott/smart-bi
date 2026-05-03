"""
Data Service - Handles data processing and transformation operations.
Follows Single Responsibility Principle - only handles data operations.
Updated to work with new file-based schema.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import polars as pl
import streamlit as st

from config import Constants, settings
from domain.entities import (
    ColumnType,
    Dashboard,
    File,
    FileSheet,
    SheetColumn,
    Visualization,
    VisualizationConfig,
    VisualizationType,
)
from domain.value_objects import AggregationConfig, FilterCondition
from infrastructure.storage.s3_client import get_s3_client

logger = logging.getLogger(__name__)


class DataService:
    """
    Service responsible for data operations including:
    - Loading and parsing XLSX files from S3
    - Schema inference and column type detection
    - Data filtering and aggregation
    - Data transformation for visualizations
    - Caching data in Streamlit session state
    """

    def __init__(self):
        """Initialize the data service."""
        # Use session state for caching
        if "data_cache" not in st.session_state:
            st.session_state.data_cache = {}
        if "sheet_cache" not in st.session_state:
            st.session_state.sheet_cache = {}

    @property
    def _data_cache(self) -> Dict[str, pl.DataFrame]:
        """Get the data cache from session state."""
        return st.session_state.data_cache

    @property
    def _sheet_cache(self) -> Dict[str, pl.DataFrame]:
        """Get the sheet cache from session state."""
        return st.session_state.sheet_cache

    def _clean_duplicate_columns(self, df: pl.DataFrame) -> pl.DataFrame:
        """
        Rename duplicate column names to make them unique.

        Args:
            df: Polars DataFrame that may have duplicate column names

        Returns:
            DataFrame with unique column names
        """
        columns = df.columns
        seen = {}
        new_columns = []

        for col in columns:
            if col in seen:
                seen[col] += 1
                new_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                new_columns.append(col)

        if new_columns != columns:
            logger.warning(
                f"Detected duplicate column names, renamed: {dict(zip(columns, new_columns))}"
            )
            df = df.rename(dict(zip(columns, new_columns)))

        return df

    def load_excel_from_streamlit(
        self, uploaded_file: Any, file_id: str, sheet_name: Optional[str] = None
    ) -> Tuple[Optional[FileSheet], Optional[pl.DataFrame]]:
        """
        Load an Excel file from Streamlit's file uploader.

        Args:
            uploaded_file: Streamlit UploadedFile object
            file_id: ID for caching
            sheet_name: Optional sheet name (uses first sheet if not specified)

        Returns:
            Tuple of (FileSheet, DataFrame), or (None, None) on failure
        """
        file_bytes = uploaded_file.getvalue()
        return self.load_excel_from_bytes(
            file_bytes, uploaded_file.name, file_id, sheet_name
        )

    def load_excel_from_bytes(
        self,
        file_bytes: bytes,
        file_name: str,
        file_id: str,
        sheet_name: Optional[str] = None,
        existing_sheet: Optional[FileSheet] = None,
    ) -> Tuple[Optional[FileSheet], Optional[pl.DataFrame]]:
        """
        Load an Excel file from bytes.

        Args:
            file_bytes: Raw bytes of the Excel file
            file_name: Original file name
            file_id: ID for caching
            sheet_name: Optional sheet name (uses first sheet if not specified)
            existing_sheet: Optional existing FileSheet entity to use its sheet_id

        Returns:
            Tuple of (FileSheet, DataFrame), or (None, None) on failure
        """
        try:
            import io

            # Get sheet names
            sheet_names = self._get_sheet_names(file_bytes)
            if not sheet_names:
                return None, None

            # Determine which sheet to read
            target_sheet = sheet_name or sheet_names[0]
            if target_sheet not in sheet_names:
                target_sheet = sheet_names[0]

            # Read the data
            df = pl.read_excel(io.BytesIO(file_bytes), sheet_name=target_sheet)
            df = self._clean_duplicate_columns(df)

            # Use existing sheet or create new one
            if existing_sheet:
                sheet = existing_sheet
                sheet_id = sheet.sheet_id
            else:
                # Create FileSheet entity with composite ID (for backward compatibility)
                sheet_id = f"{file_id}_{target_sheet}"
                sheet = FileSheet(
                    sheet_id=sheet_id,
                    file_id=file_id,
                    sheet_name=target_sheet,
                )

                # Extract columns
                columns = []
                for col_name in df.columns:
                    col_series = df[col_name]
                    polars_type = str(col_series.dtype)
                    db_type = Constants.POLARS_TO_DB_TYPE.get(polars_type, "String")

                    column = SheetColumn(
                        column_id=f"{sheet_id}_{col_name}",
                        sheet_id=sheet_id,
                        column_name=col_name,
                        data_type=db_type,
                    )
                    columns.append(column)

                sheet.columns = columns

            # Cache the DataFrame with the proper sheet_id
            st.session_state.sheet_cache[sheet_id] = df
            st.session_state.data_cache[file_id] = df

            return sheet, df

        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return None, None

    def load_file_from_s3(
        self,
        storage_path: str,
        file_id: str,
        sheet_name: Optional[str] = None,
        existing_sheet: Optional[FileSheet] = None,
    ) -> Tuple[Optional[FileSheet], Optional[pl.DataFrame]]:
        """
        Load an Excel file from S3.

        Args:
            storage_path: S3 path to the file
            file_id: ID for caching
            sheet_name: Optional sheet name (uses first sheet if not specified)
            existing_sheet: Optional existing FileSheet entity to use its sheet_id

        Returns:
            Tuple of (FileSheet, DataFrame), or (None, None) on failure
        """
        try:
            s3_client = get_s3_client()
            file_bytes = s3_client.download_file(storage_path)

            if not file_bytes:
                return None, None

            return self.load_excel_from_bytes(
                file_bytes, "file.xlsx", file_id, sheet_name, existing_sheet
            )

        except Exception as e:
            logger.error(f"Error loading file from S3: {e}")
            return None, None

    def _get_sheet_names(self, file_bytes: bytes) -> List[str]:
        """Get all sheet names from an Excel file."""
        import io
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            logger.error(f"Error getting sheet names: {e}")
            return ["Sheet1"]

    def get_cached_data(self, file_id: str) -> Optional[pl.DataFrame]:
        """Get cached DataFrame for a file."""
        return st.session_state.data_cache.get(file_id)

    def get_cached_sheet(self, sheet_id: str) -> Optional[pl.DataFrame]:
        """Get cached DataFrame for a sheet."""
        return st.session_state.sheet_cache.get(sheet_id)

    def set_cached_data(self, file_id: str, df: pl.DataFrame) -> None:
        """Set cached DataFrame for a file."""
        st.session_state.data_cache[file_id] = df

    def set_cached_sheet(self, sheet_id: str, df: pl.DataFrame) -> None:
        """Set cached DataFrame for a sheet."""
        st.session_state.sheet_cache[sheet_id] = df

    def clear_cache(
        self, file_id: Optional[str] = None, sheet_id: Optional[str] = None
    ):
        """Clear cached data."""
        if file_id:
            st.session_state.data_cache.pop(file_id, None)
        if sheet_id:
            st.session_state.sheet_cache.pop(sheet_id, None)
        if not file_id and not sheet_id:
            st.session_state.data_cache.clear()
            st.session_state.sheet_cache.clear()

    # Map of supabase data_type names -> Polars dtype for casting
    _DTYPE_MAP = {
        "Int64": pl.Int64,
        "Float64": pl.Float64,
        "String": pl.String,
        "Boolean": pl.Boolean,
        "Date": pl.Date,
        "Datetime": pl.Datetime,
        "Time": pl.Time,
    }

    def cast_column_types(
        self, df: pl.DataFrame, mapping: Dict[str, str]
    ) -> pl.DataFrame:
        """
        Cast each column in `df` to the target supabase data_type given by `mapping`.

        Args:
            df: Source DataFrame.
            mapping: {column_name: data_type_str} where data_type_str is one of
                Int64, Float64, String, Boolean, Date, Datetime, Time.

        Returns:
            New DataFrame with columns cast. Columns whose cast fails are
            converted to String as a safe fallback so the rest of the dataset
            stays usable.
        """
        result = df
        for col_name, target_type in mapping.items():
            if col_name not in result.columns:
                continue
            polars_type = self._DTYPE_MAP.get(target_type)
            if polars_type is None:
                continue
            try:
                result = result.with_columns(
                    pl.col(col_name).cast(polars_type, strict=False).alias(col_name)
                )
            except Exception as e:
                logger.warning(
                    f"Cast {col_name} -> {target_type} failed ({e}); falling back to String"
                )
                result = result.with_columns(
                    pl.col(col_name).cast(pl.String, strict=False).alias(col_name)
                )
        return result

    def apply_filters(
        self, df: pl.DataFrame, filters: List[FilterCondition]
    ) -> pl.DataFrame:
        """
        Apply filter conditions to a DataFrame.

        Args:
            df: Input DataFrame
            filters: List of filter conditions

        Returns:
            Filtered DataFrame
        """
        result = df

        for f in filters:
            col = pl.col(f.column_name)

            if f.operator == "eq":
                result = result.filter(col == f.value)
            elif f.operator == "ne":
                result = result.filter(col != f.value)
            elif f.operator == "gt":
                result = result.filter(col > f.value)
            elif f.operator == "lt":
                result = result.filter(col < f.value)
            elif f.operator == "gte":
                result = result.filter(col >= f.value)
            elif f.operator == "lte":
                result = result.filter(col <= f.value)
            elif f.operator == "contains":
                result = result.filter(col.str.contains(str(f.value)))
            elif f.operator == "starts_with":
                result = result.filter(col.str.starts_with(str(f.value)))
            elif f.operator == "ends_with":
                result = result.filter(col.str.ends_with(str(f.value)))
            elif f.operator == "in":
                result = result.filter(
                    col.is_in(f.value if isinstance(f.value, list) else [f.value])
                )
            elif f.operator == "not_in":
                result = result.filter(
                    ~col.is_in(f.value if isinstance(f.value, list) else [f.value])
                )
            elif f.operator == "is_null":
                result = result.filter(col.is_null())
            elif f.operator == "is_not_null":
                result = result.filter(col.is_not_null())

        return result

    def apply_aggregation(
        self, df: pl.DataFrame, config: AggregationConfig
    ) -> pl.DataFrame:
        """
        Apply aggregation to a DataFrame.

        Args:
            df: Input DataFrame
            config: Aggregation configuration

        Returns:
            Aggregated DataFrame
        """
        agg_col = pl.col(config.aggregation_column)

        if config.aggregation_function == "sum":
            agg_expr = agg_col.sum()
        elif config.aggregation_function == "mean":
            agg_expr = agg_col.mean()
        elif config.aggregation_function == "median":
            agg_expr = agg_col.median()
        elif config.aggregation_function == "min":
            agg_expr = agg_col.min()
        elif config.aggregation_function == "max":
            agg_expr = agg_col.max()
        elif config.aggregation_function == "count":
            agg_expr = agg_col.count()
        elif config.aggregation_function == "std":
            agg_expr = agg_col.std()
        elif config.aggregation_function == "var":
            agg_expr = agg_col.var()
        elif config.aggregation_function == "first":
            agg_expr = agg_col.first()
        elif config.aggregation_function == "last":
            agg_expr = agg_col.last()
        else:
            agg_expr = agg_col.sum()

        result = df.group_by(list(config.group_by_columns)).agg(
            agg_expr.alias(f"{config.aggregation_column}_{config.aggregation_function}")
        )

        return result

    def prepare_chart_data(
        self,
        df: pl.DataFrame,
        config: VisualizationConfig,
    ) -> pl.DataFrame:
        """
        Prepare data for chart visualization.

        Args:
            df: Input DataFrame
            config: Visualization configuration

        Returns:
            Prepared DataFrame for charting
        """
        result = df

        # Handle aggregation
        if config.x_column and config.y_column:
            if config.aggregation and config.aggregation != "none":
                agg_col = pl.col(config.y_column)

                if config.aggregation == "sum":
                    agg_expr = agg_col.sum()
                elif config.aggregation == "mean":
                    agg_expr = agg_col.mean()
                elif config.aggregation == "count":
                    agg_expr = agg_col.count()
                elif config.aggregation == "min":
                    agg_expr = agg_col.min()
                elif config.aggregation == "max":
                    agg_expr = agg_col.max()
                elif config.aggregation == "median":
                    agg_expr = agg_col.median()
                else:
                    agg_expr = agg_col.sum()

                group_cols = [config.x_column]
                if config.color_column and config.color_column != config.x_column:
                    group_cols.append(config.color_column)

                result = df.group_by(group_cols).agg(agg_expr.alias(config.y_column))

        return result

    def get_column_unique_values(
        self, df: pl.DataFrame, column_name: str, limit: Optional[int] = None
    ) -> List[Any]:
        """Get unique values from a column."""
        if limit is None:
            limit = settings.data.max_unique_values_display
        return df[column_name].unique().head(limit).to_list()

    def get_numeric_summary(
        self, df: pl.DataFrame, column_name: str
    ) -> Dict[str, float]:
        """Get summary statistics for a numeric column."""
        col = df[column_name]
        return {
            "min": float(col.min()),
            "max": float(col.max()),
            "mean": float(col.mean()),
            "median": float(col.median()),
            "std": float(col.std()),
            "q1": float(col.quantile(0.25)),
            "q3": float(col.quantile(0.75)),
        }

    def validate_and_cast_types(self, df: pl.DataFrame, schema_overrides: Dict[str, ColumnType]) -> pl.DataFrame:
        """
        Força a conversão de colunas para os tipos esperados, 
        limpando dados sujos em colunas numéricas.
        """
        for col_name, expected_type in schema_overrides.items():
            if col_name not in df.columns:
                continue
                
            if expected_type == ColumnType.NUMERIC:
                # Limpeza: remove R$, espaços e trata a vírgula como ponto
                df = df.with_columns(
                    pl.col(col_name)
                    .cast(pl.String)
                    .str.replace_all(r"[R\$\s\.]", "")
                    .str.replace(",", ".")
                    .cast(pl.Float64, strict=False)
                )
            elif expected_type == ColumnType.DATETIME:
                # Tenta converter strings para data automaticamente
                df = df.with_columns(
                    pl.col(col_name).cast(pl.Date, strict=False)
                )
            elif expected_type in (ColumnType.CATEGORICAL, ColumnType.TEXT):
                # Garante que a coluna seja string
                df = df.with_columns(
                    pl.col(col_name).cast(pl.String, strict=False)
                )

        return df

    def rename_columns(self, df: pl.DataFrame, mapping: Dict[str, str]) -> pl.DataFrame:
        """
        Renomeia colunas evitando duplicatas. 
        Se o usuário mapear duas colunas para 'Categoria', vira 'Categoria' e 'Categoria_1'.
        """
        new_names = []
        seen_names = {}

        # Criamos a lista de novos nomes verificando duplicatas
        for col in df.columns:
            if col in mapping:
                desired_name = mapping[col]
                # Se o nome já foi usado, adiciona um sufixo numérico
                if desired_name in seen_names:
                    seen_names[desired_name] += 1
                    new_name = f"{desired_name}_{seen_names[desired_name]}"
                else:
                    seen_names[desired_name] = 0
                    new_name = desired_name
                new_names.append(new_name)
            else:
                new_names.append(col) # Mantém o original se não foi mapeado

        # O Polars renomeia todas de uma vez pela ordem da lista
        return df.rename({old: new for old, new in zip(df.columns, new_names)})

    def store_data(self, analysis_id: str, df: pl.DataFrame) -> None:
        """
        Armazena o DataFrame processado no cache para ser utilizado
        pelas visualizações da análise.
        """
        st.session_state.data_cache[analysis_id] = df

    def compute_measures(self, df: pl.DataFrame, measures: list) -> pl.DataFrame:
        """
        Adiciona colunas calculadas (medidas) ao DataFrame.
        Cada medida deve ter "name" e "expression" (sintaxe [Nome da Coluna]).
        Medidas inválidas são silenciosamente ignoradas.
        """
        for measure in measures:
            name = (measure.get("name") or "").strip()
            expr_str = (measure.get("expression") or "").strip()
            if not name or not expr_str:
                continue
            try:
                expr = self._parse_measure_expr(df, expr_str)
                df = df.with_columns(expr.alias(name))
            except Exception:
                pass
        return df

    def _parse_measure_expr(self, df: pl.DataFrame, expr_str: str) -> pl.Expr:
        """
        Converte expressão com [Coluna] para Polars Expr.
        Sintaxe suportada: [Nome da Coluna] com +, -, *, / e parênteses.
        Exemplo: [Receita] / [Quantidade]
        """
        import re
        py_expr = re.sub(
            r'\[([^\]]+)\]',
            lambda m: f'pl.col("{m.group(1)}")',
            expr_str,
        )
        result = eval(py_expr, {"__builtins__": {}}, {"pl": pl})
        if not isinstance(result, pl.Expr):
            raise ValueError("Expressão deve resultar em uma coluna Polars")
        return result
