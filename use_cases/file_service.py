"""
File Service - Handles file upload, download, and processing operations.
Coordinates between storage (S3) and repository layers.
"""

from typing import Optional, Tuple, List, Dict, Any
import logging
import polars as pl
from datetime import datetime
import uuid

from config import settings, Constants
from domain.entities import File, FileSheet, SheetColumn
from infrastructure.repositories.file_repository import FileRepositoryImpl
from infrastructure.storage.s3_client import get_s3_client

logger = logging.getLogger(__name__)


class FileService:
    """
    Service responsible for file operations including:
    - Uploading files to S3
    - Downloading files from S3
    - Extracting sheet and column metadata
    - Managing file metadata in database
    """

    def __init__(self, file_repo: Optional[FileRepositoryImpl] = None):
        """
        Initialize the file service.

        Args:
            file_repo: Optional FileRepositoryImpl instance
        """
        self._file_repo = file_repo or FileRepositoryImpl()
        self._s3_client = get_s3_client()

    def upload_file(
        self,
        file_data: bytes,
        filename: str,
        user_id: str,
        content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ) -> Optional[File]:
        """
        Upload a file to S3 and save metadata to database.

        Args:
            file_data: Raw file bytes
            filename: Original filename
            user_id: User ID who owns the file
            content_type: MIME type of the file

        Returns:
            File entity with metadata, or None on failure
        """
        try:
            # Calculate file size
            file_size_kb = len(file_data) // 1024

            # Check file size limit
            if file_size_kb > settings.data.max_file_size_mb * 1024:
                logger.error(f"File too large: {file_size_kb} KB")
                return None

            # Upload to S3
            storage_path = self._s3_client.upload_file(
                file_data=file_data,
                filename=filename,
                user_id=user_id,
                content_type=content_type,
            )

            # Create file entity
            file_id = str(uuid.uuid4())
            file_entity = File(
                file_id=file_id,
                user_id=user_id,
                file_name=filename,
                storage_path=storage_path,
                file_size_kb=file_size_kb,
                uploaded_at=datetime.now(),
            )

            # Extract sheets and columns from file
            sheets = self._extract_sheets_from_bytes(file_data, file_id)
            file_entity.sheets = sheets

            # Save to database
            if self._file_repo.save_file(file_entity):
                logger.info(f"Uploaded file {file_id} for user {user_id}")
                return file_entity

            return None

        except Exception as e:
            logger.error(f"Error uploading file: {e}")
            return None

    def download_file(self, file_id: str) -> Optional[bytes]:
        """
        Download a file from S3.

        Args:
            file_id: The file ID to download

        Returns:
            File bytes, or None on failure
        """
        try:
            # Get file metadata
            file_entity = self._file_repo.find_file_by_id(file_id)
            if not file_entity:
                logger.error(f"File not found: {file_id}")
                return None

            # Download from S3
            return self._s3_client.download_file(file_entity.storage_path)

        except Exception as e:
            logger.error(f"Error downloading file {file_id}: {e}")
            return None

    def get_file(self, file_id: str) -> Optional[File]:
        """
        Get file metadata by ID.

        Args:
            file_id: The file ID

        Returns:
            File entity, or None if not found
        """
        return self._file_repo.find_file_by_id(file_id)

    def get_user_files(self, user_id: str) -> List[File]:
        """
        Get all files for a user.

        Args:
            user_id: The user ID

        Returns:
            List of File entities
        """
        return self._file_repo.find_files_by_user(user_id)

    def delete_file(self, file_id: str) -> bool:
        """
        Delete a file from S3 and database.

        Args:
            file_id: The file ID to delete

        Returns:
            True if successful, False otherwise
        """
        try:
            # Get file metadata
            file_entity = self._file_repo.find_file_by_id(file_id)
            if not file_entity:
                return False

            # Delete from S3
            self._s3_client.delete_file(file_entity.storage_path)

            # Delete from database
            return self._file_repo.delete_file(file_id)

        except Exception as e:
            logger.error(f"Error deleting file {file_id}: {e}")
            return False

    def get_sheet(self, sheet_id: str) -> Optional[FileSheet]:
        """
        Get a sheet by ID.

        Args:
            sheet_id: The sheet ID

        Returns:
            FileSheet entity, or None if not found
        """
        return self._file_repo.find_sheet_by_id(sheet_id)

    def _extract_sheets_from_bytes(
        self, file_bytes: bytes, file_id: str
    ) -> List[FileSheet]:
        """
        Extract sheets and columns from Excel file bytes.

        Args:
            file_bytes: Raw Excel file bytes
            file_id: The file ID for sheet references

        Returns:
            List of FileSheet entities with columns
        """
        import tempfile
        import os

        sheets = []

        try:
            # Write to temp file for Polars
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name

            try:
                # Read all sheets
                # Polars 0.20+ supports reading multiple sheets
                sheet_names = self._get_sheet_names(tmp_path)

                for sheet_name in sheet_names:
                    try:
                        # Read sheet data
                        df = pl.read_excel(tmp_path, sheet_name=sheet_name)

                        # Create sheet entity
                        sheet_id = str(uuid.uuid4())
                        sheet = FileSheet(
                            sheet_id=sheet_id,
                            file_id=file_id,
                            sheet_name=sheet_name,
                        )

                        # Extract columns
                        columns = self._extract_columns_from_df(df, sheet_id)
                        sheet.columns = columns

                        sheets.append(sheet)

                    except Exception as e:
                        logger.warning(f"Error reading sheet {sheet_name}: {e}")
                        continue

            finally:
                os.unlink(tmp_path)

        except Exception as e:
            logger.error(f"Error extracting sheets: {e}")

        return sheets

    def _get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get all sheet names from an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        try:
            # Use openpyxl to get sheet names
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            return wb.sheetnames
        except Exception as e:
            logger.error(f"Error getting sheet names: {e}")
            return ["Sheet1"]  # Default fallback

    def _extract_columns_from_df(
        self, df: pl.DataFrame, sheet_id: str
    ) -> List[SheetColumn]:
        """
        Extract columns from a DataFrame.

        Args:
            df: Polars DataFrame
            sheet_id: The sheet ID for column references

        Returns:
            List of SheetColumn entities
        """
        columns = []

        for col_name in df.columns:
            col_series = df[col_name]
            polars_type = str(col_series.dtype)
            db_type = Constants.POLARS_TO_DB_TYPE.get(polars_type, "String")

            column = SheetColumn(
                column_id=str(uuid.uuid4()),
                sheet_id=sheet_id,
                column_name=col_name,
                data_type=db_type,
            )
            columns.append(column)

        return columns

    def load_sheet_data(
        self, file_id: str, sheet_name: Optional[str] = None
    ) -> Tuple[Optional[pl.DataFrame], Optional[FileSheet]]:
        """
        Load data from a specific sheet.

        Args:
            file_id: The file ID
            sheet_name: Optional sheet name (uses first sheet if not specified)

        Returns:
            Tuple of (DataFrame, FileSheet), or (None, None) on failure
        """
        try:
            # Get file entity
            file_entity = self._file_repo.find_file_by_id(file_id)
            if not file_entity:
                return None, None

            # Download file
            file_bytes = self._s3_client.download_file(file_entity.storage_path)
            if not file_bytes:
                return None, None

            # Determine which sheet to read
            if sheet_name:
                sheet = next(
                    (s for s in file_entity.sheets if s.sheet_name == sheet_name),
                    file_entity.sheets[0] if file_entity.sheets else None,
                )
            else:
                sheet = file_entity.sheets[0] if file_entity.sheets else None

            if not sheet:
                return None, None

            # Read sheet data
            import tempfile
            import os

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name

            try:
                df = pl.read_excel(tmp_path, sheet_name=sheet.sheet_name)
                return df, sheet
            finally:
                os.unlink(tmp_path)

        except Exception as e:
            logger.error(f"Error loading sheet data: {e}")
            return None, None
